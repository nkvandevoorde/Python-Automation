from openpyxl import workbook
import openpyxl
from openpyxl import load_workbook
import unicodedata
import pandas as pd

#Remove special characters for consistent spelling
def remove_accents(input_str):
    if input_str is None:
        return None
    nfkd_form = unicodedata.normalize('NFKD', input_str)
    return ''.join([c for c in nfkd_form if not unicodedata.combining(c)])

#Load the Excel files
# CNES_sheet = load_workbook(filename='CNES_2024_with_REGIC_labels_p.xlsx').active
# CNES_Addresses = load_workbook(filename='cnes_output.xlsx').active

Relatorio_sheet = load_workbook(filename='relatorio-geral_COPY0510_copy.xlsx').active
Relatorio_Addresses = load_workbook(filename='relatorio_output.xlsx').active

# Concatenate number, address, and zip into a single string
#     def concatenate_address(row):
#         number = str(row[1].value) if row[1].value is not None else ''
#         address = str(row[2].value) if row[2].value is not None else ''
#         zip_code = str(row[3].value) if row[3].value is not None else ''
#         full_address = f"{number} {address} {zip_code}".strip()
#         return remove_accents(full_address).lower()

#Initiate new address column in the sheets
# num_col = CNES_sheet.max_column + 1
# CNES_sheet.cell(row=1, column=num_col).value = "Number"

# add_col = CNES_sheet.max_column + 2
# CNES_sheet.cell(row=1, column=add_col).value = "Address"

# zip_col = CNES_sheet.max_column + 3
# CNES_sheet.cell(row=1, column=zip_col).value = "Zip Code"


num_col_R = Relatorio_sheet.max_column + 1
Relatorio_sheet.cell(row=1, column=num_col_R).value = "Number"

add_col_R = Relatorio_sheet.max_column + 2
Relatorio_sheet.cell(row=1, column=add_col_R).value = "Address"

zip_col_R = Relatorio_sheet.max_column + 3
Relatorio_sheet.cell(row=1, column=zip_col_R).value = "Zip Code"

# #Create a dictionary to store the concatenated addresses from the CNES sheet
# cnes_addresses_dict = {}
# for row_idx in range(2, CNES_Addresses.max_row + 1):
#     cnes_val = str(CNES_Addresses.cell(row=row_idx, column=1).value).strip()
#     cnes_number = str(CNES_Addresses.cell(row=row_idx, column=2).value) if CNES_Addresses.cell(row=row_idx, column=2).value is not None else ''
#     cnes_address = str(CNES_Addresses.cell(row=row_idx, column=3).value) if CNES_Addresses.cell(row=row_idx, column=3).value is not None else ''
#     cnes_zip = str(CNES_Addresses.cell(row=row_idx, column=4).value) if CNES_Addresses.cell(row=row_idx, column=4).value is not None else ''
#     cnes_addresses_dict[cnes_val] = {
#         "number": cnes_number,
#         "address": cnes_address,
#         "zip": cnes_zip
#     }

# print(list(cnes_addresses_dict.items())[:5])

# #Make Address column in CNES sheet for corresponding CNES values
# matches = 0
# for row_idx in range(2, CNES_sheet.max_row + 1):
#     cnes_val = str(CNES_sheet.cell(row=row_idx, column=6).value).strip()
#     if cnes_val in cnes_addresses_dict:
#         address_info = cnes_addresses_dict[cnes_val]
#         CNES_sheet.cell(row=row_idx, column=num_col).value = address_info["number"]
#         CNES_sheet.cell(row=row_idx, column=add_col).value = address_info["address"]
#         CNES_sheet.cell(row=row_idx, column=zip_col).value = address_info["zip"]
#         matches += 1

# print(f"Number of matches in CNES sheet: {matches}")

#Update CNES excel file
# CNES_sheet.parent.save('CNES_2024_with_REGIC_labels_p.xlsx')

#Create a dictionary to store the concatenated addresses from the Relatorio sheet
relatorio_addresses_dict = {}
for row_idx in range(2, Relatorio_Addresses.max_row + 1):
    relatorio_val = str(Relatorio_Addresses.cell(row=row_idx, column=1).value).strip()
    relatorio_number = str(Relatorio_Addresses.cell(row=row_idx, column=2).value) if Relatorio_Addresses.cell(row=row_idx, column=2).value is not None else ''
    relatorio_address = str(Relatorio_Addresses.cell(row=row_idx, column=3).value) if Relatorio_Addresses.cell(row=row_idx, column=3).value is not None else ''
    relatorio_zip = str(Relatorio_Addresses.cell(row=row_idx, column=4).value) if Relatorio_Addresses.cell(row=row_idx, column=4).value is not None else ''
    relatorio_full_address = f"{relatorio_number} {relatorio_address} {relatorio_zip}".strip()
    relatorio_addresses_dict[relatorio_val] = {
        "number": relatorio_number,
        "address": relatorio_address,
        "zip": relatorio_zip
    }

print(list(relatorio_addresses_dict.items())[:5])

#Make Address column in Relatorio sheet for corresponding CNES values
matches = 0
for row_idx in range(2, Relatorio_sheet.max_row + 1):
    relatorio_val = str(Relatorio_sheet.cell(row=row_idx, column=13).value).strip()
    if relatorio_val in relatorio_addresses_dict:
        relatorio_address_info = relatorio_addresses_dict[relatorio_val]
        Relatorio_sheet.cell(row=row_idx, column=num_col_R).value = relatorio_address_info["number"]
        Relatorio_sheet.cell(row=row_idx, column=add_col_R).value = relatorio_address_info["address"]
        Relatorio_sheet.cell(row=row_idx, column=zip_col_R).value = relatorio_address_info["zip"]
        matches += 1

print(f"Number of matches in Relatorio sheet: {matches}")

#Update Relatorio excel file
Relatorio_sheet.parent.save('relatorio-geral_COPY0510_copy.xlsx')
